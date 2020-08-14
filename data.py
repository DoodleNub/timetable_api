from openpyxl import load_workbook

class component:
    pass

wb = load_workbook('timetable.xlsx')
ws = wb['Table 1']
courseCodes = []
courseSpan = []
subjectWise = {}

for row in ws.iter_rows():
    try:
        if int(row[0].value):
            courseSpan.append(int(row[0].row))
    except:
        pass

for i in range(len(courseSpan)-1):
    cell = f'B{courseSpan[i]}'
    subjectWise[f"{ws[cell].value}"] = key = []
    start = courseSpan[i]
    end = courseSpan[i+1] - 1
    for row in ws.iter_rows(min_row=start, max_row=end):
        if row[8].value:
            key.append(row[8].row)
    key.append(courseSpan[i+1])
    
objects = []

for subject in subjectWise:
    arr = subjectWise[subject]
    cell = f'I{arr[0]}' # for info regarding room number
    for i in range(len(arr)-1):
        courseObject = component()
        courseObject.subject = subject
        courseObject.teachers = []
        courseObject.room = ws[cell].value
        for row in ws.iter_rows(min_row=arr[i], max_row=arr[i+1]-1):
            courseObject.teachers.append(row[7].value)
        objects.append(courseObject)


print([(courseObject.subject, courseObject.teachers) for courseObject in objects])
    

